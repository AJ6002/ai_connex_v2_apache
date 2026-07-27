"""
aiconnex_agent/compiler_adapter.py - Adapter Wrapper between UserIntentJSON & UnifiedCompiler
=============================================================================================
Translates UserIntentJSON into UnifiedCompiler options, checks for unconfirmed HITL questions,
executes compilation, and returns a structured CompilerOutputJSON response.
"""

from __future__ import annotations

import logging
import zipfile
import openpyxl
import io
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

from aiconnex_agent.schemas import (
    UserIntentJSON,
    CompilerOutputJSON,
    HITLQuestion,
    CompilerHints,
)
from aiconnex_zip_compiler.compiler import UnifiedCompiler, CompileResult
from aiconnex_zip_compiler.plugins.registry import UnsupportedFormatError, UnsupportedLayoutError

logger = logging.getLogger(__name__)


class CompilerAdapter:
    """
    Adapter layer bridging LangGraph agent state / UserIntentJSON with UnifiedCompiler.
    """

    SUPPORTED_EXTENSIONS = {
        ".csv", ".tsv", ".txt", ".dat", ".xlsx", ".xls",
        ".parquet", ".json", ".jsonl", ".mat", ".hdf5", ".h5",
        ".tdms", ".db", ".sqlite", ".xml"
    }

    def __init__(self, output_base_dir: str = "workspace_data/compiled_output"):
        self.output_base_dir = Path(output_base_dir)
        self.output_base_dir.mkdir(parents=True, exist_ok=True)

    def analyze_zip_and_check_hitl(
        self, zip_path: Path, intent: UserIntentJSON
    ) -> List[HITLQuestion]:
        """
        Inspects dataset archive and checks if there are unconfirmed HITL questions
        that have not been answered in intent.hitl_answers.
        """
        questions: List[HITLQuestion] = []
        answers = intent.hitl_answers or {}

        if not zip_path.exists():
            return questions

        if zip_path.suffix.lower() not in [".zip", ".rar", ".7z", ".gz", ".tar"]:
            ext = zip_path.suffix.lower().lstrip(".")
            if ext not in [s.lstrip(".") for s in self.SUPPORTED_EXTENSIONS]:
                raise UnsupportedFormatError(ext)
            return questions

        # Inspect Zip contents
        try:
            with zipfile.ZipFile(zip_path, "r") as z:
                namelist = z.namelist()
                xlsx_files = [f for f in namelist if f.endswith(".xlsx") or f.endswith(".xls")]
                
                # Check format validity of contained files
                valid_exts = {s.lstrip(".") for s in self.SUPPORTED_EXTENSIONS}.union({"txt", "md", "json", "doc", "pdf"})
                for f in namelist:
                    if not f.endswith("/") and not Path(f).name.startswith(".") and not f.startswith("__MACOSX"):
                        f_ext = Path(f).suffix.lower().lstrip(".")
                        if f_ext and f_ext not in valid_exts:
                            raise UnsupportedFormatError(f_ext)

                if xlsx_files:
                    first_xlsx = xlsx_files[0]
                    data = z.read(first_xlsx)
                    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
                    sheetnames = wb.sheetnames

                    # Q1: Multi-sheet selection
                    if len(sheetnames) > 1 and "Q1_sheet_selection" not in answers:
                        questions.append(
                            HITLQuestion(
                                key="Q1_sheet_selection",
                                question=f"The Excel file(s) contain multiple sheets: {sheetnames}. Which sheet(s) should be compiled?",
                                options=[f"{s} only" for s in sheetnames] + ["Both/All — join on Date"],
                                blocking=True,
                                reason="Multiple sheets serve different analytical goals."
                            )
                        )

                    # Q2: Multi-row header collapsing
                    ws = wb[sheetnames[0]]
                    rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
                    merged_header_detected = False
                    for r in rows[:5]:
                        non_none = [x for x in r if x is not None]
                        if len(non_none) > 0 and any("report" in str(x).lower() or "station" in str(x).lower() for x in non_none):
                            merged_header_detected = True
                            break

                    if merged_header_detected and "Q2_header_row_depth" not in answers:
                        questions.append(
                            HITLQuestion(
                                key="Q2_header_row_depth",
                                question="The Excel file has multi-row metadata headers at the top. How should column headers be parsed?",
                                options=[
                                    "Auto-detect and collapse header rows",
                                    "Use only row 9 (first data header)",
                                    "Concatenate top 3 header rows"
                                ],
                                blocking=True,
                                reason="Multi-row headers require clear collapsing rules."
                            )
                        )

                    # Q3: Target selection if ambiguous
                    if not intent.target_column and "Q6_target_column" not in answers:
                        questions.append(
                            HITLQuestion(
                                key="Q6_target_column",
                                question="Which target column or metric would you like to predict/analyze?",
                                options=[
                                    "Auto-detect target column",
                                    "Gas loss / gain (in Kg)",
                                    "Total Gas Inlet Per Day",
                                    "Total Gas Outlet Per Day"
                                ],
                                blocking=False,
                                reason="Target column specification guides feature engineering and algorithm selection."
                            )
                        )

        except zipfile.BadZipFile:
            raise ValueError(f"File '{zip_path}' is not a valid zip archive.")

        return questions

    def execute(self, zip_path: str | Path, intent: UserIntentJSON) -> CompilerOutputJSON:
        """
        Executes compiler pipeline or returns HITL questions if clarification is needed.
        """
        zip_p = Path(zip_path)
        output_dir = self.output_base_dir / intent.session_id

        try:
            # 1. Check for format issues or HITL requirements
            hitl_questions = self.analyze_zip_and_check_hitl(zip_p, intent)
            blocking_questions = [q for q in hitl_questions if q.blocking]

            if blocking_questions:
                logger.info(f"[CompilerAdapter] HITL clarification required ({len(hitl_questions)} questions)")
                return CompilerOutputJSON(
                    session_id=intent.session_id,
                    status="hitl_required",
                    hitl_required=True,
                    hitl_questions=hitl_questions,
                    warnings=["User clarification required before compilation can proceed."]
                )

            # 2. Invoke UnifiedCompiler
            compiler = UnifiedCompiler(
                zip_path=zip_p,
                output_dir=output_dir,
                enable_intelligence=False,
                interactive=False
            )
            compile_res: CompileResult = compiler.compile()

            if not compile_res.success:
                return CompilerOutputJSON(
                    session_id=intent.session_id,
                    status="failed",
                    hitl_required=False,
                    error=compile_res.error or "Compilation failed without detailed error trace."
                )

            # 3. Extract schema from compiled CSV
            target_csv = compile_res.combined_file or (compile_res.merged_files[0] if compile_res.merged_files else None)
            detected_schema = {}
            row_count = 0
            col_count = 0

            if target_csv and Path(target_csv).exists():
                import pandas as pd
                df_sample = pd.read_csv(target_csv, nrows=100)
                detected_schema = {col: str(dtype) for col, dtype in df_sample.dtypes.items()}
                col_count = len(df_sample.columns)
                # Count total rows safely
                with open(target_csv, "r", encoding="utf-8", errors="ignore") as f:
                    row_count = max(0, sum(1 for _ in f) - 1)

            return CompilerOutputJSON(
                session_id=intent.session_id,
                status="success",
                compiled_csv_path=str(target_csv) if target_csv else None,
                row_count=row_count,
                column_count=col_count,
                detected_schema=detected_schema,
                hitl_required=False,
                compiler_decisions={
                    "duration_seconds": compile_res.duration_seconds,
                    "merged_files_count": len(compile_res.merged_files),
                    "combined_file": compile_res.combined_file,
                },
                warnings=[]
            )

        except UnsupportedFormatError as e:
            logger.warning(f"[CompilerAdapter] Unsupported format: {e}", exc_info=True)
            return CompilerOutputJSON(
                session_id=intent.session_id,
                status="unsupported_format",
                hitl_required=False,
                error=str(e)
            )
        except Exception as e:
            logger.error(f"[CompilerAdapter] Execution error: {e}", exc_info=True)
            return CompilerOutputJSON(
                session_id=intent.session_id,
                status="failed",
                hitl_required=False,
                error=f"Compiler exception: {str(e)}"
            )
