"""
XLSX Parser Worker - Defusedxml-safe Excel sheet processing & region extraction to Parquet.
Executes inside parser-xlsx sandbox container under non-root 10001:10001 with read-only rootfs.
"""

import hashlib
import os
import sys
from datetime import datetime
from pathlib import Path

import defusedxml
import openpyxl
import polars as pl
import pyarrow.parquet as pq

from contracts.sandbox.result_manifest_contract import ParserResultManifest

# Enable defusedxml protections globally for XML parsing in openpyxl
defusedxml.defuse_stdlib()


def compute_sha256(file_path: Path) -> str:
    h = hashlib.sha256()
    with open(file_path, "rb") as f:
        while chunk := f.read(65536):
            h.update(chunk)
    return h.hexdigest()


def process_xlsx():
    start_time = datetime.utcnow()
    input_dir = Path(os.environ.get("SANDBOX_INPUT_DIR", "/sandbox/input"))
    output_dir = Path(os.environ.get("SANDBOX_OUTPUT_DIR", "/sandbox/output"))
    output_dir.mkdir(parents=True, exist_ok=True)

    input_files = list(input_dir.glob("*.xlsx")) + list(input_dir.glob("*.xls"))
    if not input_files:
        print(f"No XLSX input files found in {input_dir}")
        sys.exit(1)

    input_path = input_files[0]
    input_hash = compute_sha256(input_path)

    output_parquet_path = output_dir / f"{input_path.stem}.parquet"

    try:
        # Open workbook in data_only mode safely
        wb = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
        sheet_names = wb.sheetnames
        first_sheet = wb[sheet_names[0]]

        data_rows = []
        for row in first_sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                data_rows.append(list(row))

        wb.close()

        if not data_rows:
            print("XLSX file is empty", file=sys.stderr)
            sys.exit(1)

        # Extract header and rows
        header = [str(col) if col is not None else f"col_{idx}" for idx, col in enumerate(data_rows[0])]
        rows = data_rows[1:]

        # Create Polars DataFrame -> Arrow Table
        df = pl.DataFrame(rows, schema=header, orient="row")
        arrow_table = df.to_arrow()
        row_count = arrow_table.num_rows

        pq.write_table(arrow_table, output_parquet_path, compression="snappy")
        output_hash = compute_sha256(output_parquet_path)

        schema_def = {col: str(dtype) for col, dtype in zip(arrow_table.schema.names, arrow_table.schema.types)}

        end_time = datetime.utcnow()

        manifest = ParserResultManifest(
            job_id=os.environ.get("JOB_ID", f"job_xlsx_{int(start_time.timestamp())}"),
            image_name="parser-xlsx",
            image_digest=os.environ.get("IMAGE_DIGEST", "sha256:local_unpinned"),
            input_file=input_path.name,
            input_hash=input_hash,
            output_parquet=output_parquet_path.name,
            output_hash=output_hash,
            row_count=row_count,
            schema_definition=schema_def,
            started_at=start_time,
            completed_at=end_time,
            lineage={
                "parser": "parser-xlsx",
                "engine": "openpyxl+defusedxml+polars",
                "format": "xlsx",
                "sheet": sheet_names[0]
            }
        )

        manifest_path = output_dir / "result_manifest.json"
        with open(manifest_path, "w", encoding="utf-8") as f:
            f.write(manifest.model_dump_json(indent=2))

        print(f"XLSX Parser successfully generated Parquet artifact: {output_parquet_path} ({row_count} rows)")

    except Exception as e:
        print(f"XLSX Parser Error: {e!s}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    process_xlsx()
